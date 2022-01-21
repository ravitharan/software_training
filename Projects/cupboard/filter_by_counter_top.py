import openpyxl # THIS MODULE TO HANDLE xl FILES
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
import sys
import re

TEMPLATE_FILE = "ct_template.xlsx"

def xl_filter(order_wb):
    active_sheet = order_wb["Main Sheet"]
    template = openpyxl.load_workbook(filename=TEMPLATE_FILE)
    template_sheet = template["Main Sheet"]
    maximum_row = active_sheet.max_row
    maximum_column = active_sheet.max_column

    template_row  = 6
    ct_count      = 0

    for row in range(1, maximum_row + 1):
        """Header write"""
        if isinstance(active_sheet.cell(row, 1).value, str):
            if  "DATE ISSUE:" in active_sheet.cell(row, 1).value:
                header_columns = ['C', 'H', 'L', 'M', 'O', 'X', 'AC', 'AG', 'AN']
                for col_name in header_columns:
                    column = column_index_from_string(col_name)
                    if col_name != 'AN':
                        template_sheet.cell(row, column).value = active_sheet.cell(row, column).value
                    else:
                        template_sheet.cell(row+1, column).value = active_sheet.cell(row+1, column).value
        """Write counter top existing row"""
        if bool(re.search(r'\d', str(active_sheet.cell(row, column_index_from_string('AD')).value))):
            for row_read in range(row , row+3):
                for column in range(1, maximum_column+1):
                    if active_sheet.cell(row_read, column).value != None:
                        template_sheet.cell(template_row, column).value = active_sheet.cell(row_read, column).value
                template_row += 1
            ct_count += 1

            if (ct_count == 8):
                ct_count = 0
                template_row += 5

    return template

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print(f'Argument error\nUsage: {sys.argv[0]} <user_excel_file>')
        exit(1)
    user_xl_file = openpyxl.load_workbook(filename= sys.argv[1], data_only=True)
    wb_out = xl_filter(user_xl_file)
    wb_out.save("ct_output.xlsx")
