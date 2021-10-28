from parse_vanity_info import *
from parse_order import *
from openpyxl.styles import *

SHEET_STYLE_COUNT       = "STYLE COUNT"
SHEET_COLOR_COUNT       = "COLOR COUNT"
SHEET_STYLE_COLOR_COUNT = "STYLE COLOR COUNT"

START_ROW = 3

def adjust_column_width(work_sheets):
    for ws in work_sheets:
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            if length > 20:
                length = 20
            ws.column_dimensions[column_cells[0].column_letter].width = length+5


def write_parts_for_workshop(work_sheets, items):
    '''
    Seperate all items by material 
    '''
    materials = [ x[1] for x in items ]
    materials = list(set(materials))
    materials.sort()

    start_row = START_ROW
    for material in materials:
        items_for_material = [x for x in items if x[1] == material]
        start_row += write_material_items(work_sheets, start_row, items_for_material)


def write_material_items(work_sheets, start_row, items):
    '''
    Write material & style header and seperate items by part name
    '''

    row = start_row
    for ws in work_sheets:
        ws.cell(row, 1).value = "MATERIAL"
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = items[0][1]
        ws.cell(row, 2).font = Font(bold=True)
        ws.cell(row, 2).alignment = Alignment(horizontal="center")
    row += 1

    styles = [ x[2] for x in items ]
    styles = list(set(styles))
    styles.sort()

    colors = [ x[3] for x in items ]
    colors = list(set(colors))
    colors.sort()

    for ws in work_sheets:
        ws.cell(row, 1).value = "ITEM"
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 1).alignment = Alignment(horizontal="center")
        ws.cell(row, 2).value = "SIZE"
        ws.cell(row, 2).font = Font(bold=True)
        ws.cell(row, 2).alignment = Alignment(horizontal="center")

        if ws.title == SHEET_COLOR_COUNT:
            for i, color in enumerate(colors):
                ws.cell(row, i+3).value = color
                ws.cell(row, i+3).font = Font(bold=True)
                ws.cell(row, i+3).alignment = Alignment(horizontal="center")
        else:
            for i, style in enumerate(styles):
                ws.cell(row, i+3).value = style
                ws.cell(row, i+3).font = Font(bold=True)
                ws.cell(row, i+3).alignment = Alignment(horizontal="center")
    row += 1

    parts = [ x[5] for x in items ]
    parts = list(set(parts))
    parts.sort()

    #Re-arrange the parts order
    if 'DOOR' in parts:
        parts.remove('DOOR')
        parts.insert(0,'DOOR')
    if 'DRAWER' in parts:
        parts.remove('DRAWER')
        parts.insert(1,'DRAWER')

    for part in parts:
        part_items = [x for x in items if x[5] == part]
        #print(f'PART: {part}')
        details, sizes = get_part_details(part_items, styles, colors)
        #print(details)
        #print()
        row = write_part_names(work_sheets, row, part, details, sizes, styles, colors)
        row += 3

    return row

def get_part_details(items, styles, colors):
    '''
    input list as [[count, material, style, color, size, part_name], ..]
    Get details for a specific part as
    details = 
    {
     '8 7/8 X 26 1/2': {'AHM 80': 2, 'VISTA FLAT': [2, '2 - AHM 80'], 'total': [2, 2]},
     '11 7/8 X 26 1/2': {'AHM 20 MATTE': 2, 'AHM 50': 2, 'SHAKER': [2, '2 - AHM 50'], 'SIERRA FLAT': [2, '2 - AHM 20 MATTE'], 'total': [4, 4]},
     '11 7/8 X 17 9/16': {'AHM 10 MATTE': 4, 'VISTA FLAT': [4, '4 - AHM 10 MATTE'], 'total': [4, 4]},
     '11 7/8 X 19 1/2': {'AHM 10 MATTE': 2, 'SIERRA RAISED': [2, '2 - AHM 10 MATTE'], 'total': [2, 2]},
     '14 7/8 X 17 9/16': {'AHM 10 MATTE': 4, 'AHM 40': 4, 'VISTA FLAT': [8, '4 - AHM 10 MATTE\n4 - AHM 40'], 'total': [8, 8]},
     '14 7/8 X 44 1/8': {'AHM 10 MATTE': 1, 'VISTA FLAT': [1, '1 - AHM 10 MATTE'], 'total': [1, 1]},
     '14 7/8 X 26 1/2': {'AHM 10 MATTE': 1, 'AHM 25': 2, 'VISTA FLAT': [3, '1 - AHM 10 MATTE\n2 - AHM 25'], 'total': [3, 3]}
    }
    '''
    details = {}

    sizes = [ x[4] for x in items ]
    sizes = list(set(sizes))
    sizes = sorted(sizes, key = lambda x: int(x.split()[0]))

    for size in sizes:
        details[size] = {}
        total_colors = 0
        total_styles = 0
        for color in colors:
            count = [x[0] for x in items if x[4] == size and x[3] == color ]
            if sum(count):
                details[size][color] = sum(count)
            total_colors += sum(count)
        for style in styles:
            style_items = [x for x in items if x[4] == size and x[2] == style ]
            color_per_style = []
            for color in colors:
                color_count = [x[0] for x in style_items if x[3] == color ]
                if sum(color_count):
                    color_per_style.append(f'{sum(color_count)} - {color}')

            count = [x[0] for x in items if x[4] == size and x[2] == style ]
            if sum(count):
                details[size][style] = [sum(count), '\n'.join(color_per_style) ]
            total_styles += sum(count)
        details[size]['total'] = [total_styles, total_colors]
    return details, sizes


def write_part_names(work_sheets, start_row, item, details, sizes, styles, colors):
    '''
    Write detail for a specific part. Rows contain sizes and columns contain styles
    '''
    #excel cell styles
    border_style = Border(top = Side(style='double'), bottom = Side(style='thin'))

    row = start_row

    for ws in work_sheets:
        ws.cell(row,1).value = item
        ws.cell(row, 1).font = Font(bold=True)

    styles_total = [0] * len(styles)
    colors_total  = [0] * len(colors)

    for size in sizes:
        if size in details:
            for ws in work_sheets:
                ws.cell(row,2).value = size

                if ws.title == SHEET_STYLE_COUNT:
                    for i, style in enumerate(styles):
                        if style in details[size]:
                            ws.cell(row, i+3).value = details[size][style][0]
                            ws.cell(row, i+3).alignment = Alignment(horizontal="center")
                            styles_total[i] += details[size][style][0]

                elif ws.title == SHEET_STYLE_COLOR_COUNT:
                    for i, style in enumerate(styles):
                        if style in details[size]:
                            ws.cell(row,i+3).value = details[size][style][1]
                            ws.cell(row,i+3).alignment = Alignment(wrap_text=True)

                elif ws.title == SHEET_COLOR_COUNT:
                    for i, color in enumerate(colors):
                        if color in details[size]:
                            ws.cell(row, i+3).value = details[size][color]
                            ws.cell(row, i+3).alignment = Alignment(horizontal="center")
                            colors_total[i] += details[size][color]
        row += 1

    #total of parts
    for ws in work_sheets:
        ws.cell(row,1).value = 'Total'
        ws.cell(row,2).border = border_style
        ws.cell(row,2).alignment = Alignment(horizontal="center")
        if ws.title == SHEET_COLOR_COUNT:
            ws.cell(row,2).value = sum(colors_total)
            for i, color in enumerate(colors):
                ws.cell(row,i+3).border = border_style
                ws.cell(row,i+3).alignment = Alignment(horizontal="center")
                if colors_total[i]:
                    ws.cell(row,i+3).value = colors_total[i]
        else:
            ws.cell(row,2).value = sum(styles_total)
            for i, style in enumerate(styles):
                ws.cell(row,i+3).border = border_style
                ws.cell(row,i+3).alignment = Alignment(horizontal="center")
                if styles_total[i]:
                    ws.cell(row,i+3).value = styles_total[i]

    return row+1

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print(f'Argument error\nUsage: {sys.argv[0]} <cupboard_excel_file>')
        exit(1)

    excel_file = sys.argv[1]
    wb = openpyxl.load_workbook(excel_file)
    cupboard_parts = get_cupboard_list(wb)
    items, work_week = get_order_list(wb, cupboard_parts)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_STYLE_COUNT
    wb.create_sheet(SHEET_STYLE_COLOR_COUNT)
    wb.create_sheet(SHEET_COLOR_COUNT)

    ws = [ wb[SHEET_STYLE_COUNT], wb[SHEET_STYLE_COLOR_COUNT], wb[SHEET_COLOR_COUNT] ]

    write_parts_for_workshop(ws, items)
    adjust_column_width(ws)

    wb.save('output.xlsx')

