import re
import sys
import openpyxl
from openpyxl.styles import *
from openpyxl.utils import get_column_letter

SHEET_COUNTER_TOP   = "COUNTER TOP"
RIGHT_SPLASH        = 'RIGHT SPLASH'
LEFT_SPLASH         = 'LEFT SPLASH'

def get_counter_tops(wb):
    '''
    Get valid cupboard list from column C and D in "VANITY INFO" sheet.
    Return dictionary of cupboard entries.
    '''
    ws = wb["Main Sheet"]
    maximum_row = ws.max_row
    counter_tops = []
    counter_tops_sink = []
    size_sink = []
    sizes = set()
    colors = set()
    sinks = set()
    sink_colors = set()
    
    rect_count = 0
    oval_count = 0
    for x in range(1 , maximum_row + 1):
        order_number  = ws["AD" + str(x)].value 
        size  = ws["AF" + str(x)].value
        color = ws["AG" + str(x)].value
        sink  = ws["AH" + str(x)].value
        sink_splash = ws["Z" + str(x)].value
        if (type(order_number) == int) or bool(re.search(r'\d',str(order_number))):
            if size != None:
                size  = size.strip()
                if color != None:
                    color = color.strip()
                if sink != None:
                    sink  = sink.strip()
                size = size.replace('"', '')
                if sink == "RECTANGULAR":
                    rect_count += 1
                elif sink == "OVAL":
                    oval_count += 1
                else:
                    print(f'Invalid sink type "{sink}" at cell AH{x}')
                counter_tops.append([size, color])
                sizes.add(size)
                colors.add(color)

                if sink_splash == "R" or sink_splash == "LR":
                    counter_tops_sink.append([size,color,RIGHT_SPLASH])
                    size_sink.append([f'{size}_{RIGHT_SPLASH}'])
                    sink_colors.add(color)
                    #counter_tops.append([size, RIGHT_SPLASH])
                if sink_splash == "L" or sink_splash == "LR":
                    counter_tops_sink.append([size, color, LEFT_SPLASH])
                    size_sink.append([f'{size}_{LEFT_SPLASH}'])
                    sink_colors.add(color)
                    #counter_tops.append([size, LEFT_SPLASH])
                 
    sizes = list(sizes)
    colors = list(colors)
    sinks = list(sinks)
    sink_colors = list(sink_colors)
    
    #sizes.sort(key = lambda x: int(x.split()[0]))
    
    colors.sort() 
    #colors.append('')
    sinks.append(RIGHT_SPLASH)
    sinks.append(LEFT_SPLASH)
     
    return counter_tops, list(sizes), list(colors), [rect_count, oval_count], counter_tops_sink, size_sink, sink_colors 

def get_countertop_details(counter_tops, sizes, colors, counter_tops_sink,size_sink):

    details = {}
     
    for item in counter_tops:
         
        if (item[0] in details) and (item[1] in details[item[0]]):
            details[item[0]][item[1]] += 1
        else:
            if item[0] not in details:
                details[item[0]] = {}
            details[item[0]][item[1]] = 1

    size_sink_set = set()
    for List in size_sink:
        size_sink_set.add(List[0])

    sink_details = {}
    for data in size_sink_set:
        data = data.split('_')
        for List in counter_tops_sink:
            if data[0] == List[0] and data[1] == List[2]:
                key = List[0]+ '-' + List[2]
                if (key in sink_details):
                    if (List[1] in sink_details[key]):
                        sink_details[key][List[1]] += 1
                    else:  
                        sink_details[key][List[1]] = 1    
                else:
                    sink_details[key] = {}
                    sink_details[key][List[1]] = 1
    
    return details, sink_details

def write_counter_top(ws, start_row, details, sizes, colors, sink_counts, sink_colors, sink_details):

    row = start_row
    max_col_width = [0] * (1 + len(colors))

    for i, color in enumerate(colors):
        max_col_width[i+1] = len(color)
        ws.cell(row, i+2).value = color
        ws.cell(row, i+2).font = Font(bold=True)
        ws.cell(row, i+2).alignment = Alignment(horizontal="center")
    row += 1

    sink_splash_total = 0
    total = [0] * len(colors)
    for size in sizes:
        max_col_width[0] = max(max_col_width[0], len(size))
        ws.cell(row, 1).value = size
        for i, color in enumerate(colors):
            if (size in details) and (color in details[size]):
                ws.cell(row, i+2).value = details[size][color]
                ws.cell(row, i+2).alignment = Alignment(horizontal="center")
                total[i] += details[size][color]
                if color == RIGHT_SPLASH or color == LEFT_SPLASH:
                    sink_splash_total += details[size][color]
        row += 1

    border_style = Border(top = Side(style='double'), bottom = Side(style='thin'))
    ws.cell(row, 1).value = sum(total) - sink_splash_total
    ws.cell(row, 1).alignment = Alignment(horizontal="center")
    ws.cell(row, 1).border = border_style
    for i, color in enumerate(colors):
        if total[i]:
            ws.cell(row, i+2).value = total[i]
            ws.cell(row, i+2).alignment = Alignment(horizontal="center")
            ws.cell(row, i+2).border = border_style

    row += 2

    for i, color in enumerate(sink_colors):
        max_col_width[i+1] = len(color)
        ws.cell(row, i+2).value = color
        ws.cell(row, i+2).font = Font(bold=True)
        ws.cell(row, i+2).alignment = Alignment(horizontal="center")
    row += 1
    
    sink_total = [0] * len(sink_colors)
    for size_sink in sink_details:
        max_col_width[0] = max(max_col_width[0], len(size))
        ws.cell(row, 1).value = size_sink
        for i, color in enumerate(sink_colors):
            if (color in sink_details[size_sink]):    
                ws.cell(row, i+2).value = (sink_details[size_sink][color])
                ws.cell(row, i+2).alignment = Alignment(horizontal="center")
                sink_total[i] += sink_details[size_sink][color]
                 
        row += 1

    border_style = Border(top = Side(style='double'), bottom = Side(style='thin'))
    ws.cell(row, 1).value = sum(sink_total)
    ws.cell(row, 1).alignment = Alignment(horizontal="center")
    ws.cell(row, 1).border = border_style

    for i, color in enumerate(sink_colors):
        if sink_total[i]:
            ws.cell(row, i+2).value = sink_total[i]
            ws.cell(row, i+2).alignment = Alignment(horizontal="center")
            ws.cell(row, i+2).border = border_style
    row += 2
    
    ws.cell(row, 1).value = "RECTANGULAR"
    ws.cell(row, 1).font = Font(bold=True)
    ws.cell(row, 2).value = sink_counts[0]
    ws.cell(row, 2).alignment = Alignment(horizontal="center")
    row += 1
    
    ws.cell(row, 1).value = "OVAL"
    ws.cell(row, 1).font = Font(bold=True)
    ws.cell(row, 2).value = sink_counts[1]
    ws.cell(row, 2).alignment = Alignment(horizontal="center")
    
    for i, col_width in enumerate(max_col_width):
        ws.column_dimensions[get_column_letter(i+1)].width = (col_width + 2)* 1.9
            
    return row

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print(f'Argument error\nUsage: {sys.argv[0]} <cupboard_excel_file>')
        exit(1)

    excel_file = sys.argv[1]
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    counter_tops, sizes, colors, counts, counter_tops_sink, size_sink, sink_colors = get_counter_tops(wb)
    details, sink_details =  get_countertop_details(counter_tops, sizes, colors, counter_tops_sink, size_sink)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_COUNTER_TOP

    write_counter_top(ws, 3, details, sizes, colors, counts, sink_colors, sink_details)

    wb.save('output.xlsx')
