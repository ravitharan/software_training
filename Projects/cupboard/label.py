from openpyxl import load_workbook, Workbook, drawing
from PIL import Image
from openpyxl_image_loader import SheetImageLoader
from openpyxl.styles import *
from openpyxl.utils import *
import re
import sys

BIG_LABEL_TEMPLATE = "big_label_template.xlsx"

def get_vanity_info_cupboard_name(wb):
    """
    Get cupboard  ID and Parts from column A and B in "VANITY INFO" sheet.Return dictionary of entries

    """
    wb_sheet = wb["VANITY INFO"]
    maximum_row = wb_sheet.max_row
    cupboard_id_parts = {}
    for x in range(1 , maximum_row + 1):
        value_A = wb_sheet["A" + str(x)].value
        value_B = wb_sheet["B" + str(x)].value
        if (type(value_A) == int):
            value_B = re.sub(" +"," ", value_B)
            cupboard_id_parts[value_A] = value_B

    return cupboard_id_parts

def get_vanity_info_valance_detail(wb):
    """

    """
    wb_sheet = wb["VANITY INFO"]
    maximum_row = wb_sheet.max_row
    valance_detail = {}
    for x in range(1 , maximum_row + 1):
        value_R = wb_sheet["R" + str(x)].value
        value_S = wb_sheet["S" + str(x)].value
        if value_R and value_S:
            valance_detail[value_R] = value_S

    return valance_detail

def add_row(cell_name, row_offset):
    (col, row) = cell.coordinate_from_string(cell_name)
    return col + str(row + row_offset)

def cupboard_data_write(wb_r, ws, cupboard_names):
    """
    Main sheet and vanity info data write in to tha Label tamplate
    
    """
    ws_r = wb_r["Main Sheet"]
    label_count = 0
    # row_label is as 1, 18, 33, 50, 65, 82, 97, 114, 129, 146, 161, 178, 193, 210
    row_label = 0
    label_cell_map = [
            ('H1', 'C6'),
            ('H2', 'B8'),
            ('H3', 'D6'),
            ('G4', 'H1'),
            ('A7', 'A6'),
            ('A13','K8'),
            ('B9', 'F6'),
            ('B13','K6'),
            ('F9', 'H6'),
            ('F10','G6'),
            ('F11','I6'),
            ('F12','L6'),
            ('F13','K3'),
            ('G12','L7'),
            ]

    for row_order in range(1, ws_r.max_row+1):
        order_no  = ws_r["C" + str(row_order)].value
        cupboard_id  = ws_r["E" + str(row_order)].value
        if (isinstance(cupboard_id, int) or (('w' in str(cupboard_id)) or ('W' in str(cupboard_id)))) and (bool(re.search(r'\d',str(order_no)))) :
            label_count += 1
            if (label_count == 1):
                row_label = 1
            elif ((label_count % 2) == 0):
                row_label += 17
            else:
                row_label += 15
                
            #image insertion
            img = drawing.image.Image('logo_small.png')
#            img.height=130
#            img.width=200
            img.anchor = 'A'+str(row_label)
            ws.add_image(img)

            for (dst, src) in label_cell_map:
                if dst == 'G4':
                    dst_cell = add_row(dst, row_label - 1)
                    src_cell = ws_r["H1"].value
                    ws[dst_cell].value = src_cell
                else:
                    dst_cell = add_row(dst, row_label - 1)
                    src_cell = add_row(src, row_order - 6)
                    #print(f'dst {dst_cell}, src {src_cell}')
                    ws[dst_cell].value = ws_r[src_cell].value

            if ('w' in str(cupboard_id)) or ('W' in str(cupboard_id)):
                continue
            
            else:
                dst_cell = add_row('B10', row_label - 1)
                cupboard_info = cupboard_names[cupboard_id]
                ws[dst_cell].value = cupboard_info

                dst_cell = add_row('A9', row_label - 1)
                ws[dst_cell].value = '[1]'

     

def framed_mirror_data_write(wb_r, ws):
    """
    Main sheet data write in to tha Label tamplate
    
    """
    ws_r = wb_r["Main Sheet"]
    label_count = 0
    # row_label is as 1, 18, 33, 50, 65, 82, 97, 114, 129, 146, 161, 178, 193, 210
    row_label = 0
    framed_label_cell_map = [
                            ('Q1', 'C6'),
                            ('Q2', 'B8'),
                            ('Q3', 'D6'),
                            ('P4', 'H1'),
                            ('J7', 'A6'),
                            ('J9', 'M8'),
                            ('K9', 'M6'),
                            ('K10','M3'),
                            ('O9', 'H6'),
                            ('O10','G6'),
                            ('O11','I6'),
                    
                            ]

    for row_order in range(1, ws_r.max_row+1):
        order_no  = ws_r["C" + str(row_order)].value
        framed_mirror  = ws_r["M" + str(row_order)].value
        if (type(order_no) == str or type(order_no) == int) and framed_mirror and bool(re.search(r'\d',str(order_no))):
            label_count += 1
            if (label_count == 1):
                row_label = 1
            elif ((label_count % 2) == 0):
                row_label += 17
            else:
                row_label += 15
                
            #image insertion
            img = drawing.image.Image('logo_small.png')
#            img.height=130
#            img.width=200
            img.anchor = 'J'+str(row_label)
            ws.add_image(img)

            for (dst, src) in framed_label_cell_map:
                if dst == 'P4':
                    dst_cell = add_row(dst, row_label - 1)
                    src_cell = ws_r["H1"].value
                    ws[dst_cell].value = src_cell
                elif dst == 'K10':
                    dst_cell = add_row(dst, row_label - 1)
                    src_cell = ws_r["M3"].value
                    ws[dst_cell].value = src_cell
                else:
                    dst_cell = add_row(dst, row_label - 1)
                    src_cell = add_row(src, row_order - 6)
                    #print(f'dst {dst_cell}, src {src_cell}')
                    ws[dst_cell].value = ws_r[src_cell].value
     

def valance_data_write(wb_r, ws, valances):
    """
    Main sheet data write in to tha Label tamplate
    
    """
    ws_r = wb_r["Main Sheet"]
    label_count = 0
    # row_label is as 1, 18, 33, 50, 65, 82, 97, 114, 129, 146, 161, 178, 193, 210
    row_label = 0
    valance_label_cell_map = [
                            ('Z1', 'C6'),
                            ('Z2', 'B8'),
                            ('Z3', 'D6'),
                            ('Y4', 'H1'),
                            ('S7', 'A6'),
                            ('T9', 'J6'),
                            ('T10','J3'),
                            ('X9', 'H6'),
                            ('X10','G6'),
                            ('X11','I6'),
                    
                            ]

    for row_order in range(1, ws_r.max_row+1):
        order_no = ws_r["E" + str(row_order)].value
        valance  = ws_r["J" + str(row_order)].value
        if (type(order_no) == str or type(order_no) == int) and valance and bool(re.search(r'\d',str(order_no))):
            label_count += 1
            if (label_count == 1):
                row_label = 1
            elif ((label_count % 2) == 0):
                row_label += 17
            else:
                row_label += 15
                
            #image insertion
            img = drawing.image.Image('logo_small.png')
#            img.height=130
#            img.width=200
            img.anchor = 'S'+str(row_label)
            ws.add_image(img)

            for (dst, src) in valance_label_cell_map:
                if dst == 'Y4':
                    dst_cell = add_row(dst, row_label - 1)
                    src_cell = ws_r["H1"].value
                    ws[dst_cell].value = src_cell
                else:
                    dst_cell = add_row(dst, row_label - 1)
                    src_cell = add_row(src, row_order - 6)
                    #print(f'dst {dst_cell}, src {src_cell}')
                    ws[dst_cell].value = ws_r[src_cell].value

            dst_cell = add_row('S9', row_label - 1)
            ws[dst_cell].value = '[1]'

            dst_cell = add_row('T12', row_label - 1)
            valance_code  = ws_r["J" + str(row_order+2)].value
            valance_val = valances[valance_code]
            ws[dst_cell].value = valance_val

def counter_top_data_write(wb_r, ws_template):
    """
    Main sheet data write in to tha Label tamplate
    
    """
    ws_r = wb_r["Main Sheet"]
    ws_w = ws_template
    label_count = 0
    # row_label is as 1, 18, 33, 50, 65, 82, 97, 114, 129, 146, 161, 178, 193, 210
    row_label = 0
    counter_top_label_cell_map = [
                            ('AI1', 'AD6'),
                            ('AI2', 'B8'),
                            ('AI4', 'D6'),
                            ('AI5', 'H1'),
                            ('AB7', 'A6'),
                            ('AG9','AG3'),
                            ('AC9', 'AF6'),
                            ('AH9','AG6'),
                            ('AB10','Y6'),
                            ('AC10','Y3'),
                            ('AG10','AH3'),
                            ('AH10','AH6'),
                            ('AH11','AI6'),
                            ('AG11','AI3'),
                            ('AH12','AK6'),
                            ('AG12','AK3'),
                            ('AG13','AJ3'),
                            ('AH13','AJ6'),
                            ('AC13','AA6')
                            ]

    for row_order in range(1, ws_r.max_row+1):
        order_number  = ws_r["AD" + str(row_order)].value  
        if (type(order_number) == int) or bool(re.search(r'\d',str(order_number))):
            label_count += 1
            if (label_count == 1):
                row_label = 1
            elif ((label_count % 2) == 0):
                row_label += 17
            else:
                row_label += 15
                
            #image insertion
            img = drawing.image.Image('logo_small.png')
                #img.height=130
                #img.width=200
            img.anchor = 'AB'+str(1+row_label)
            ws_w.add_image(img)
            
            add_list =[('AC10','Y3'),('AI5', 'H1'),('AG9','AG3'),('AG10','AH3'),('AG11','AI3'),('AG12','AK3'),('AG13','AJ3')]
            for (dst, src) in counter_top_label_cell_map:
                for (r,w) in add_list:
                    if dst == r:
                        dst_cell = add_row(dst, row_label - 1)
                        src_cell = ws_r[w].value
                        ws_w[dst_cell].value = src_cell
                        break
                 
                    else:
                        dst_cell = add_row(dst, row_label - 1)
                        src_cell = add_row(src, row_order - 6)
                        #print(f'dst {dst_cell}, src {src_cell}')
                        ws_w[dst_cell].value = ws_r[src_cell].value
     
def write_big_labels(wb):
    wb_template = load_workbook(BIG_LABEL_TEMPLATE)
    ws_template = wb_template.active
     
    cupboard_names = get_vanity_info_cupboard_name(wb)
    cupboard_data_write(wb, ws_template, cupboard_names)

    framed_mirror_data_write(wb, ws_template)
    valances = get_vanity_info_valance_detail(wb)
    valance_data_write(wb, ws_template, valances)
    counter_top_data_write(wb, ws_template)
    return wb_template

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print(f'Argument error\nUsage: {sys.argv[0]} <order_file>')
        exit(1)
    wb_order = load_workbook(filename= sys.argv[1], data_only=True)
    wb_out = write_big_labels(wb_order)
    wb_out.save("big_label.xlsx")
