import re
import sys
import openpyxl
from openpyxl.styles import *
from openpyxl.utils import get_column_letter

SHEET_COUNTER_TOP = "COUNTER TOP"
Right_Splash = 'Right_Splash'
Left_Splash = 'Left_Splash'
def get_counter_tops(wb):
    '''
    Get valid cupboard list from column C and D in "VANITY INFO" sheet.
    Return dictionary of cupboard entries.
    '''
    ws = wb["Main Sheet"]
    maximum_row = ws.max_row
    counter_tops = []
    sizes = set()
    colors = set()
    
    rect_count = 0
    oval_count = 0
    for x in range(1 , maximum_row + 1):
        cupboard_id  = ws["E" + str(x)].value
        size  = ws["AF" + str(x)].value
        color = ws["AG" + str(x)].value
        sink  = ws["AH" + str(x)].value
        sink_splash = ws["Z" + str(x)].value
        if isinstance(cupboard_id, int) and size and color and sink:
            size  = size.strip()
            color = color.strip()
            sink  = sink.strip()
            size = size.replace('"', '')
            if sink == "RECTANGULAR":
                rect_count += 1
            elif sink == "OVAL":
                oval_count += 1
            else:
                print(f'Invalid sink type "{sink}" at cell AH{x}')
            counter_tops.append([size, color])
            sizes.add(size)
            colors.add(color)
            
            if sink_splash == "R" or sink_splash == "LR":
                counter_tops.append([size, Right_Splash])
                colors.add(Right_Splash)
            if sink_splash == "L" or sink_splash == "LR":
                counter_tops.append([size, Left_Splash])
                colors.add(Left_Splash)
            
    sizes = list(sizes)
    colors = list(colors)

    sizes.sort(key = lambda x: int(x.split()[0]))
    colors.sort()
     
    if Right_Splash in colors:
        colors.remove(Right_Splash)
        colors.append(Right_Splash)
    if Left_Splash in colors:
        colors.remove(Left_Splash)
        colors.append(Left_Splash)
    
    return counter_tops, list(sizes), list(colors), [rect_count, oval_count]

def get_countertop_details(counter_tops, sizes, colors):

    details = {}

    for item in counter_tops:
        if (item[0] in details) and (item[1] in details[item[0]]):
            details[item[0]][item[1]] += 1
        else:
            if item[0] not in details:
                details[item[0]] = {}
            details[item[0]][item[1]] = 1
    
    return details

def write_counter_top(ws, start_row, details, sizes, colors, sink_counts):

    row = start_row
    max_col_width = [0] * (1 + len(colors))
     
    for i, color in enumerate(colors):
        max_col_width[i+1] = len(color)
        if color == Right_Splash or color == Left_Splash:
            ws.cell(row, i+3).value = color
            ws.cell(row, i+3).font = Font(bold=True)
            ws.cell(row, i+3).alignment = Alignment(horizontal="center")
        else:
            ws.cell(row, i+2).value = color
            ws.cell(row, i+2).font = Font(bold=True)
            ws.cell(row, i+2).alignment = Alignment(horizontal="center")
    row += 1
     
    sink_splash_total = 0
    total = [0] * len(colors)
    for size in sizes:
        max_col_width[0] = max(max_col_width[0], len(size))
        ws.cell(row, 1).value = size
        for i, color in enumerate(colors):
            if (size in details) and (color in details[size]):
                if color == Right_Splash or color == Left_Splash:
                    ws.cell(row, i+3).value = details[size][color]
                    ws.cell(row, i+3).alignment = Alignment(horizontal="center")
                    total[i] += details[size][color]
                    sink_splash_total += details[size][color]
                else:
                    ws.cell(row, i+2).value = details[size][color]
                    ws.cell(row, i+2).alignment = Alignment(horizontal="center")
                    total[i] += details[size][color]
        row += 1
    
    border_style = Border(top = Side(style='double'), bottom = Side(style='thin'))
    ws.cell(row, 1).value = sum(total) - sink_splash_total
    ws.cell(row, 1).alignment = Alignment(horizontal="center")
    ws.cell(row, 1).border = border_style
    for i, color in enumerate(colors):
        if total[i]:
            if color == Right_Splash or color == Left_Splash:
                ws.cell(row, i+3).value = total[i]
                ws.cell(row, i+3).alignment = Alignment(horizontal="center")
                ws.cell(row, i+3).border = border_style
            else:
                ws.cell(row,i+2).value = total[i]
                ws.cell(row, i+2).alignment = Alignment(horizontal="center")
                ws.cell(row, i+2).border = border_style
    
    for i, col_width in enumerate(max_col_width):
        ws.column_dimensions[get_column_letter(i+2)].width = col_width * 1.5

    row += 2
    ws.cell(row, 1).value = "RECTANGULAR"
    ws.cell(row, 1).font = Font(bold=True)

    ws.cell(row, 2).value = sink_counts[0]
    ws.cell(row, 2).alignment = Alignment(horizontal="center")

    row += 1
    ws.cell(row, 1).value = "OVAL"
    ws.cell(row, 1).font = Font(bold=True)

    ws.cell(row, 2).value = sink_counts[1]
    ws.cell(row, 2).alignment = Alignment(horizontal="center")
    return row

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print(f'Argument error\nUsage: {sys.argv[0]} <cupboard_excel_file>')
        exit(1)

    excel_file = sys.argv[1]
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    counter_tops, sizes, colors, counts = get_counter_tops(wb)
    details =  get_countertop_details(counter_tops, sizes, colors)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_COUNTER_TOP

    write_counter_top(ws, 3, details, sizes, colors, counts)

    wb.save('output.xlsx')
