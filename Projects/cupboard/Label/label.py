from openpyxl import load_workbook, Workbook, drawing
from PIL import Image
from openpyxl_image_loader import SheetImageLoader
from openpyxl.styles import *
from openpyxl.utils import *
import re

wb = load_workbook("cupboard_parts.xlsx", data_only=True)

def get_vanity_info_data(wb):
    """
    Get cupboard  ID and Parts from column A and B in "VANITY INFO" sheet.Return dictionary of entries

    """
    wb_sheet = wb["VANITY INFO"]
    maximum_row = wb_sheet.max_row
    cupboard_id_parts = {}
    for x in range(1 , maximum_row + 1):
        value_A = wb_sheet["A" + str(x)].value
        value_B = wb_sheet["B" + str(x)].value
        if (type(value_A) == int) and (value_A != None) :
            value_B = re.sub(" +"," ", value_B)
            cupboard_id_parts[value_A] = value_B

    return cupboard_id_parts

cupboard_id_parts = get_vanity_info_data(wb)

wb_r = load_workbook("WORK.xlsx", data_only=True)
ws_r = wb_r["Main Sheet"]
wb_w = load_workbook("template.xlsx")
ws_w = wb_w.active
label_count = 0
# row_label is as 1, 18, 33, 50, 65, 82, 97, 114, 129, 146, 161, 178, 193, 210
row_label = 0
label_cell_map = [
        ('H1', 'C6'),
        ('H2', 'B8'),
        ('H3', 'D6'),
        ('G4', 'H1'),
        ('A7', 'A6'),
        ('B9', 'F6'),
        ('B10','Dic'),
        ('F9', 'H6'),
        ('F10','G6'),
        ('F11','I6'),
        ('F12','L6'),
        ('G12','L7'),
        ('F13','K3'),
        ('B13','K6'),
        ('A13','K8')
        ]

def add_row(cell_name, row_offset):
    (col, row) = cell.coordinate_from_string(cell_name)
    return col + str(row + row_offset)

for row_order in range(1, ws_r.max_row+1):
    cupboard_id  = ws_r["E" + str(row_order)].value
    if isinstance(cupboard_id, int):
        label_count += 1
        if (label_count == 1):
            row_label = 1
        elif ((label_count % 2) == 0):
            row_label += 17
        else:
            row_label += 15
        #image insertion

        img = drawing.image.Image('logo.png')
        img.height=130
        img.width=200
        img.anchor = 'A'+str(1+row_label)
        ws_w.add_image(img)

        for (dst, src) in label_cell_map:
            if dst == 'B10':
                dst_cell = add_row(dst, row_label - 1)
                src_cell = cupboard_id_parts[cupboard_id]
                ws_w[dst_cell].value = src_cell
            else:
                dst_cell = add_row(dst, row_label - 1)
                src_cell = add_row(src, row_order - 6)
                #print(f'dst {dst_cell}, src {src_cell}')
                ws_w[dst_cell].value = ws_r[src_cell].value
            
wb_w.save("big_label.xlsx")
