import openpyxl  # A python module to handle Excel files
import re        #  Regular Expressions (RegEx) is a special sequence of characters that uses a search pattern to find a string or set of strings
import sys

def clean_description(description):
    '''
    clean parts description
    '''
    description = description.replace("x" , "X")
    description = description.replace("-" , " ")
    description =  re.sub(' +', ' ', description)
    return description

def split_parts_details(parts_description):
    '''
    This will split a parts description (column D value) into [[count1, size1, name1], [count2, size2, name2], ...]
    Input string should have been cleaned
    eg '[2] 9 7/8 X 26 1/2 DOORS [1] 11 7/8 X 6 15/16 DRAWER [1] 11 7/8 X 8 15/16 DRAWER [1] 11 7/8 X 10 7/16 DRAWER [2] 2 X 32 SOLID WOOD [4] 2 X 2 X 32 SOLID WOOD LEG'
    Above description become 
        [
            [2, '9 7/8 X 26 1/2', 'DOORS'],
            [1, '11 7/8 X 6 15/16', 'DRAWER'],
            [1, '11 7/8 X 8 15/16', 'DRAWER'],
            [1, '11 7/8 X 10 7/16', 'DRAWER'],
            [2, '2 X 32', 'SOLID WOOD'],
            [4, '2 X 2 X 32', 'SOLID WOOD LEG']
        ]

    '''
    # Make sure parts_description contain at least one [n]
    if not re.compile('\[\d+\]').search(parts_description):
        print(f'INVALID parts: {parts_description}')
        return []

    # Split string by count ('[n]') entries
    part_entries = re.compile('(\[\d+\])').split(parts_description)
    while '' in part_entries:               # Remove blank entry
        part_entries.remove('')

    parts = []

    for i in range(0, len(part_entries), 2):
        count = int(part_entries[i][1:-1])    # get the number by skipping [ & ]
        size_name = part_entries[i+1]
        # Split string into size and name ('DOOR') entries
        size_name_entries = re.compile('([a-wyzA-WYZ].+)').split(size_name)
        while '' in size_name_entries:               # Remove blank entry
            size_name_entries.remove('')
        if size_name_entries[1].strip() != 'CLASSIC KICK':
            if size_name_entries[1].strip() == 'MOULDING':
                parts.append([count, size_name_entries[0].strip(), size_name_entries[1].strip()+" & CLASSIC KICK"])
            else:
                parts.append([count, size_name_entries[0].strip(), size_name_entries[1].strip()])
         
    return parts

def get_cupboard_list(wb):
    '''
    Get valid cupboard list from column C and D in "VANITY INFO" sheet.
    Return dictionary of cupboard entries.

    '''
    ws = wb["VANITY INFO"]
    maximum_row = ws.max_row
    cupboard_parts = {}
    for x in range(1 , maximum_row + 1):
        value_C = ws["C" + str(x)].value
        value_D = ws["D" + str(x)].value
        if (type(value_C) == int) and (value_D != None) :
            description = clean_description(value_D)
            parts_list = split_parts_details(description)
            cupboard_parts[value_C] = parts_list
        
    return cupboard_parts

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print(f'Argument error\nUsage: {sys.argv[0]} <cupboard_excel_file>')
        exit(1)

    excel_file = sys.argv[1]
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    cupboard_parts = get_cupboard_list(wb)
    """for key in cupboard_parts:
        print(key , " ")
        for parts in cupboard_parts[key]:
            print(parts)"""

