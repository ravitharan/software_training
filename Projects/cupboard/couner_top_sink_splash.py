import re
import sys
import openpyxl
from openpyxl.styles import *
from openpyxl.utils import get_column_letter

SHEET_COUNTER_TOP = "COUNTER TOP"

def get_counter_tops(wb):
    '''
    Get valid cupboard list from column C and D in "VANITY INFO" sheet.
    Return dictionary of cupboard entries.
    '''
    ws = wb["Main Sheet"]
    maximum_row = ws.max_row
    counter_tops = []
    sinks_splash_list=[]
    sizes = set()
    colors = set()
    sinks_splash = set()

    rect_count = 0
    oval_count = 0
    for x in range(1 , maximum_row + 1):
        cupboard_id  = ws["E" + str(x)].value
        size  = ws["AF" + str(x)].value
        color = ws["AG" + str(x)].value
        sink  = ws["AH" + str(x)].value
        sink_splash = ws["Z" + str(x)].value
        if isinstance(cupboard_id, int) and size and color and sink:
            size  = size.strip()
            color = color.strip()
            sink  = sink.strip()
            sink_splash = sink_splash.strip()
            size = size.replace('"', '')
            if sink == "RECTANGULAR":
                rect_count += 1
            elif sink == "OVAL":
                oval_count += 1
            else:
                print(f'Invalid sink type "{sink}" at cell AH{x}')
            counter_tops.append([size, color])
            sinks_splash_list.append([size, sink_splash])
            sizes.add(size)
            colors.add(color)
            sinks_splash.add(sink_splash)

    sizes = list(sizes)
    colors = list(colors)
    sinks_splash = list(sinks_splash)

    sizes.sort(key = lambda x: int(x.split()[0]))
    colors.sort()
    sinks_splash.sort()
    
    return counter_tops, sinks_splash_list, list(sizes), list(colors), list(sinks_splash), [rect_count, oval_count]

def get_countertop_details(counter_tops, sizes, colors):

    details = {}
    
    for item in counter_tops:
        if (item[0] in details) and (item[1] in details[item[0]]):
            details[item[0]][item[1]] += 1
        else:
            if item[0] not in details:
                details[item[0]] = {}
            details[item[0]][item[1]] = 1

    return details

def Sink_Splash(sinks_splash_list):
    Sink_Splash = {}

    for value in sinks_splash_list:
        if (value[0] in Sink_Splash) and (value[1] in Sink_Splash[value[0]]):
            Sink_Splash[value[0]][value[1]] += 1
        else:
            if value[0] not in Sink_Splash:
                Sink_Splash[value[0]] = {}
            Sink_Splash[value[0]][value[1]] = 1
            
    return Sink_Splash

def write_counter_top(ws, start_row, details, sizes, colors, sink_counts):
    
    row = start_row
    max_col_width = [0] * (1 + len(colors))

    for i, color in enumerate(colors):
        max_col_width[i+1] = len(color)
        ws.cell(row, i+2).value = color
        ws.cell(row, i+2).font = Font(bold=True)
        ws.cell(row, i+2).alignment = Alignment(horizontal="center")

    i += 3
    for n, sink_splash in enumerate(sinks_splash):
        if sink_splash == "LR":
                continue
        else:
            if sink_splash == "R":
                ws.cell(row, i+1).value = "Right" 
                ws.cell(row, i+1).font = Font(bold=True)
                ws.cell(row, i+1).alignment = Alignment(horizontal="center")
            else:
                if sink_splash == "L":
                    ws.cell(row, i+1).value = "Left"
                    ws.cell(row, i+1).font = Font(bold=True)
                    ws.cell(row, i+1).alignment = Alignment(horizontal="center")
            i += 1
            
    row += 1
    str_row = row
    st_row = row - 1
    total = [0] * len(colors)
    for size in sizes:
        max_col_width[0] = max(max_col_width[0], len(size))
        ws.cell(row, 1).value = size
        for i, color in enumerate(colors):
            if (size in details) and (color in details[size]):
                ws.cell(row, i+2).value = details[size][color]
                ws.cell(row, i+2).alignment = Alignment(horizontal="center")
                total[i] += details[size][color]
        row += 1
        column = i + 4
    tot1,tot2 = 0, 0   
    for key in Sink_Splash:
        for k, value in enumerate(Sink_Splash[key]):
            if k == 0:
                st_row += 1
            if value == "R":
                ws.cell(st_row, column+1).value = Sink_Splash[key][value]
                tot1 += Sink_Splash[key][value]
            elif value == "L":
                ws.cell(st_row, column).value = Sink_Splash[key][value]
                tot2 += Sink_Splash[key][value]
            else:
                if value == "LR":
                    ws.cell(st_row, column+1).value = Sink_Splash[key][value]
                    tot1 += Sink_Splash[key][value]
                    ws.cell(st_row, column).value = Sink_Splash[key][value]
                    tot2 += Sink_Splash[key][value]
    
    border_style = Border(top = Side(style='double'), bottom = Side(style='thin'))
    ws.cell(row, 1).value = sum(total)
    ws.cell(row, 1).alignment = Alignment(horizontal="center")
    ws.cell(row, 1).border = border_style
    
    for i, color in enumerate(colors):
        if total[i]:
            ws.cell(row, i+2).value = total[i]
            ws.cell(row, i+2).alignment = Alignment(horizontal="center")
            ws.cell(row, i+2).border = border_style
            
    ws.cell(st_row+1, column).value = tot2
    ws.cell(st_row+1, column).border = border_style
    ws.cell(st_row+1, column+1).value = tot1
    ws.cell(st_row+1, column+1).border = border_style
    
    for i, col_width in enumerate(max_col_width):
        ws.column_dimensions[get_column_letter(i+1)].width = col_width * 1.5

    row += 2
    ws.cell(row, 1).value = "RECTANGULAR"
    ws.cell(row, 1).font = Font(bold=True)

    ws.cell(row, 2).value = sink_counts[0]
    ws.cell(row, 2).alignment = Alignment(horizontal="center")

    row += 1
    ws.cell(row, 1).value = "OVAL"
    ws.cell(row, 1).font = Font(bold=True)

    ws.cell(row, 2).value = sink_counts[1]
    ws.cell(row, 2).alignment = Alignment(horizontal="center")
    return row

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print(f'Argument error\nUsage: {sys.argv[0]} <cupboard_excel_file>')
        exit(1)

    excel_file = sys.argv[1]
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    counter_tops, sinks_splash_list, sizes, colors, sinks_splash, counts = get_counter_tops(wb)
    details =  get_countertop_details(counter_tops, sizes, colors)
    Sink_Splash = Sink_Splash(sinks_splash_list)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_COUNTER_TOP

    write_counter_top(ws, 3, details, sizes, colors, counts)

    wb.save('output.xlsx')
