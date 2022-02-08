import re
import sys
import openpyxl
import logging
from openpyxl.styles import *
from openpyxl.utils import get_column_letter

SHEET_KNOB_COUNT        = "KNOB COUNT"
CUPBOARD_ID_COLUMN      = "E"
HARDWARE_ITEM_COLUMN    = "L"
ORDER_NUMBER            = "C"

def hardware_data(wb):
    current_sheet = wb["Main Sheet"]
    maximum_row = current_sheet.max_row
    logging.info(f'Main Sheet : HARDWARE DATA')
    hardware_count_items = []
    for x in range(1 , maximum_row + 1):
        cupboard_id = current_sheet[CUPBOARD_ID_COLUMN + str(x)].value
        order_no  = current_sheet[ORDER_NUMBER + str(x)].value

        # creat log file for finding errors
        logging.info(f'{"  "}ROW NUMBER :{x} ,{"  "} Cell Address -{CUPBOARD_ID_COLUMN + str(x)} : {cupboard_id}, {"  "}, Cell Address -{ORDER_NUMBER + str(x)} : {order_no}')

        if (isinstance(cupboard_id, int) or (('w' in str(cupboard_id)) or ('W' in str(cupboard_id)))) and (bool(re.search(r'\d',str(order_no)))) :
            count_items = []
            c = x + 2
            hardware_count  = current_sheet[HARDWARE_ITEM_COLUMN + str(c)].value
            if hardware_count != None:
                hardware_count = re.findall('\d+', hardware_count)
                for count in hardware_count:
                    count_items.append(int(count))
                    
                hardware_door = current_sheet[HARDWARE_ITEM_COLUMN + str(x)].value
                #Remove "DR-"
                if hardware_door:
                    hardware_door = hardware_door[3:]
                     
                hardware_drawer = current_sheet[HARDWARE_ITEM_COLUMN + str(x+1)].value
                 
                #Remove "DW-"
                if hardware_drawer:
                    hardware_drawer = hardware_drawer[3:]

                if hardware_door:
                    hardware_count_items.append([count_items[0], hardware_door])
                    
                if len(count_items) == 2:
                    hardware_count_items.append([count_items[1], hardware_drawer]) 
                     
    return hardware_count_items
 
def hardware_add_count(hardware_count_item):
    hardware_item_set = set()
    for item_list in hardware_count_item:
        hardware_item_set.add(item_list[1])

    hardware_count_Dict = {}
    for set_item in hardware_item_set:
        total = 0
        for list_item in hardware_count_item:
            if list_item[1] == set_item:
                total += list_item[0]
        hardware_count_Dict[set_item] = total
     
    return hardware_count_Dict

def Hardware_Details_Write(ws, start_row, hardware_count_Dict_data):
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 24
    
    row = start_row
    ws.cell(row, 1).value = 'Hardware Item' # Excel_cell
    ws.cell(row, 1).font = Font(bold=True, size = 16)
#    ws.cell(row, 1).fill = PatternFill(fgColor="61E43A", fill_type = "solid")
    ws.cell(row, 1).alignment = Alignment(horizontal='center')
    
    ws.cell(row, 2).value = 'Count' # Excel_cell
    ws.cell(row, 2).font = Font(bold=True, size = 16)
#    ws.cell(row, 2).fill = PatternFill(fgColor="61E43A", fill_type = "solid")
    ws.cell(row, 2).alignment = Alignment(horizontal='center')

    row += 1
    total = 0
    for key in hardware_count_Dict_data:
        ws["A" + str(row)] = key # Excel_cell_value_add
        ws["B" + str(row)] = hardware_count_Dict_data[key] # Excel_cell_value_add
        ws["B"+ str(row)].alignment = Alignment(horizontal='center')
        total += ws["B" + str(row)].value
        row += 1
    
    ws["A" + str(row+1)] = "Total"
    ws["A" + str(row+1)].font = Font(bold=True, size = 12)
    
    ws["B" + str(row+1)] = total
    ws["B" + str(row+1)].font = Font(bold=True, size = 12)
    ws["B"+ str(row+1)].alignment = Alignment(horizontal='center')
    
    return row

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print(f'Argument error\nUsage: {sys.argv[0]} <cupboard_excel_file>')
        exit(1)

    excel_file = sys.argv[1]    
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    hardware_count_items = hardware_data(wb)
    hardware_count_Dict = hardware_add_count(hardware_count_items)   

    excel_file = openpyxl.Workbook()
    sheet = excel_file.active
    sheet.title = SHEET_KNOB_COUNT
    #sheet.title = "Hardware_Details"
    Hardware_Details_Write(sheet, 3, hardware_count_Dict)

    excel_file.save('Hardware_Details.xlsx') # New Excel file create
