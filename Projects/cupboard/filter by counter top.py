import openpyxl # THIS MODULE TO HANDLE EXCEL FILES
from openpyxl.utils import get_column_letter

user_excel_file = openpyxl.load_workbook(filename = "C:\\Users\\A2Z Lankan\\Desktop\\sithu\\cupboard_parts.xlsx" , data_only = True)
template = openpyxl.load_workbook(filename = "C:\\Users\\A2Z Lankan\\Desktop\\sithu\\ct_template.xlsx")
active_sheet = user_excel_file["Main Sheet"]
template_sheet = template["Sheet1"]
maximum_row = active_sheet.max_row
maximum_column = active_sheet.max_column
template_sheet["C1"] = active_sheet["C1"].value
template_sheet["C30"] = active_sheet["C30"].value
template_sheet["H1"] = active_sheet["H1"].value
template_sheet["H30"] = active_sheet["H30"].value
template_sheet["X1"] = active_sheet["X1"].value
template_sheet["X30"] = active_sheet["X30"].value
template_sheet["AC1"] = active_sheet["AC1"].value
template_sheet["AC30"] = active_sheet["AC30"].value
template_sheet["AG1"] = active_sheet["AG1"].value
template_sheet["AG30"] = active_sheet["AG30"].value
for row in range(6 , maximum_row+1):
    if active_sheet["AD" + str(row)].value != None:
        if "COMPANY" in active_sheet[get_column_letter(1) + str(row)].value:
            print(row)
            template_sheet[get_column_letter(1) + str(row)] = active_sheet[get_column_letter(1) + str(row)].value
            template_sheet[get_column_letter(1) + str(row+2)] = active_sheet[get_column_letter(1) + str(row+2)].value
            template_sheet[get_column_letter(2) + str(row+2)] = active_sheet[get_column_letter(2) + str(row+2)].value
            for col in range(3,10):
                template_sheet[get_column_letter(col) + str(row)] = active_sheet[get_column_letter(col) + str(row)].value
            for col in range(10,12):
                template_sheet[get_column_letter(col) + str(row)] = active_sheet[get_column_letter(col) + str(row)].value
                template_sheet[get_column_letter(col) + str(row+2)] = active_sheet[get_column_letter(col) + str(row+2)].value
            template_sheet[get_column_letter(12) + str(row)] = active_sheet[get_column_letter(12) + str(row)].value
            template_sheet[get_column_letter(12) + str(row+1)] = active_sheet[get_column_letter(12) + str(row+1)].value
            template_sheet[get_column_letter(12) + str(row+2)] = active_sheet[get_column_letter(12) + str(row+2)].value
            template_sheet[get_column_letter(13) + str(row)] = active_sheet[get_column_letter(13) + str(row)].value
            template_sheet[get_column_letter(13) + str(row+2)] = active_sheet[get_column_letter(13) + str(row+2)].value
            for col in range(14,27):
                template_sheet[get_column_letter(col) + str(row)] = active_sheet[get_column_letter(col) + str(row)].value
            template_sheet[get_column_letter(27) + str(row)] = active_sheet[get_column_letter(27) + str(row)].value
            template_sheet[get_column_letter(28) + str(row)] = active_sheet[get_column_letter(28) + str(row)].value
            template_sheet[get_column_letter(28) + str(row+2)] = active_sheet[get_column_letter(28) + str(row+2)].value
            template_sheet[get_column_letter(29) + str(row+2)] = active_sheet[get_column_letter(29) + str(row+2)].value
            for col in range(30,41):
                template_sheet[get_column_letter(col) + str(row)] = active_sheet[get_column_letter(col) + str(row)].value

template.save("C:\\Users\\A2Z Lankan\\Desktop\\sithu\\ct_template.xlsx")      
        
        

    

