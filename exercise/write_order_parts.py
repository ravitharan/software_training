import sys
import openpyxl

def write_parts_for_workshop(work_sheets, items):
    '''
    Seperate all items by material 
    '''
    materials = [ x[1] for x in items ]
    materials = list(set(materials))
    materials.sort()

    start_row = 1
    for material in materials:
        items_for_material = [x for x in items if x[1] == material]
        start_row += write_material_items(work_sheets, start_row, items_for_material)


def write_material_items(work_sheets, start_row, items):
    '''
    Write material & style header and seperate items by part name
    '''

    row = start_row
    for ws in work_sheets:
        ws.cell(row, 1).value = "MATERIAL"
        ws.cell(row, 2).value = items[0][1]
    row += 1

    styles = [ x[2] for x in items ]
    styles = list(set(styles))
    styles.sort()

    for ws in work_sheets:
        ws.cell(row, 1).value = "ITEM"
        ws.cell(row, 2).value = "SIZE"

        for i, style in enumerate(styles):
            ws.cell(row, i+3).value = style
    row += 1

    part_names = [ x[5] for x in items ]
    part_names = list(set(part_names))
    part_names.sort()

    for part_name in part_names:
        part_items = [x for x in items if x[5] == part_name]
        row = write_part_names(work_sheets, row, part_items, styles)
        row += 1

    return row

def write_part_names(work_sheets, start_row, items, styles):
    '''
    Write detail for a specific part. Rows contain sizes and columns contain styles
    '''
    sizes = [ x[4] for x in items ]
    sizes = list(set(sizes))
    sizes = sorted(sizes, key = lambda x: int(x.split()[0]))

    row = start_row

    '''
    Need to write rest of the code
    '''

    row += 1

    return row

if __name__ == "__main__":

    #Use other modules to generate items list as below
    items = [
        [1, 'MAPLE', 'CAPRICE FLAT', 'NATURAL', '41 7/8 X 4', 'CLASSIC KICK'],
        [2, 'MAPLE', 'SHAKER', 'AHM 3700', '47 7/8 X 4', 'CLASSIC KICK'],
        [1, 'MAPLE', 'SHAKER', 'AHM 3700', '59 7/8 X 4', 'CLASSIC KICK'],
        [8, 'MAPLE', 'SHAKER', 'AHM 3700', '11 7/8 X 26 1/2', 'DOOR'],
        [2, 'MAPLE', 'SHAKER', 'AHM 3700', '11 7/8 X 31 7/8', 'DOOR'],
        [2, 'MAPLE', 'CAPRICE FLAT', 'NATURAL', '11 7/8 X 31 7/8', 'DOOR'],
        [2, 'MAPLE', 'CAPRICE FLAT', 'NATURAL', '8 7/8 X 26 1/2', 'DOOR'],
        [15, 'MAPLE', 'SHAKER', 'AHM 3700', '11 7/8 X 8 3/4', 'DRAWER'],
        [6, 'MAPLE', 'CAPRICE FLAT', 'NATURAL', '11 7/8 X 8 3/4', 'DRAWER'],
        [1, 'MAPLE', 'CAPRICE FLAT', 'NATURAL', '41 7/8 X 2 3/4', 'MOULDING'],
        [2, 'MAPLE', 'SHAKER', 'AHM 3700', '47 7/8 X 2 3/4', 'MOULDING'],
        [1, 'MAPLE', 'SHAKER', 'AHM 3700', '59 7/8 X 2 3/4', 'MOULDING'],
        [1, 'MDF', 'VISTA FLAT', 'AHM 10 MATTE`', '14 7/8 X 4', 'CLASSIC KICK'],
        [1, 'MDF', 'SIERRA FLAT', 'AHM 20 MATTE', '23 7/8 X 4', 'CLASSIC KICK'],
        [1, 'MDF', 'VISTA FLAT', 'AHM 25', '29 7/8 X 4', 'CLASSIC KICK'],
        [1, 'MDF', 'SHAKER', 'AHM 50', '35 7/8 X 4', 'CLASSIC KICK'],
        [1, 'MDF', 'VISTA FLAT', 'AHM 80', '41 7/8 X 4', 'CLASSIC KICK'],
        [1, 'MDF', 'VISTA FLAT', 'AHM  10 MATTE', '59 7/8 X 4', 'CLASSIC KICK'],
        [1, 'MDF', 'VISTA FLAT', 'AHM 10 MATTE', '71 7/8 X 4', 'CLASSIC KICK'],
        [1, 'MDF', 'VISTA FLAT', 'AHM 40', '71 7/8 X 4', 'CLASSIC KICK'],
        [4, 'MDF', 'VISTA FLAT', 'AHM  10 MATTE', '11 7/8 X 17 9/16', 'DOOR'],
        [2, 'MDF', 'SIERRA RAISED', 'AHM 10 MATTE', '11 7/8 X 19 1/2', 'DOOR'],
        [2, 'MDF', 'SHAKER', 'AHM 50', '11 7/8 X 26 1/2', 'DOOR'],
        [2, 'MDF', 'SIERRA FLAT', 'AHM 20 MATTE', '11 7/8 X 26 1/2', 'DOOR'],
        [4, 'MDF', 'VISTA FLAT', 'AHM 10 MATTE', '14 7/8 X 17 9/16', 'DOOR'],
        [4, 'MDF', 'VISTA FLAT', 'AHM 40', '14 7/8 X 17 9/16', 'DOOR'],
        [2, 'MDF', 'VISTA FLAT', 'AHM 25', '14 7/8 X 26 1/2', 'DOOR'],
        [1, 'MDF', 'VISTA FLAT', 'AHM 10 MATTE`', '14 7/8 X 26 1/2', 'DOOR'],
        [1, 'MDF', 'VISTA FLAT', 'AHM 10 MATTE`', '14 7/8 X 44 1/8', 'DOOR'],
        [2, 'MDF', 'VISTA FLAT', 'AHM 80', '8 7/8 X 26 1/2', 'DOOR'],
        [3, 'MDF', 'SHAKER', 'AHM 50', '11 7/8 X 8 3/4', 'DRAWER'],
        [3, 'MDF', 'VISTA FLAT', 'AHM  10 MATTE', '11 7/8 X 8 3/4', 'DRAWER'],
        [6, 'MDF', 'VISTA FLAT', 'AHM 80', '11 7/8 X 8 3/4', 'DRAWER'],
        [3, 'MDF', 'VISTA FLAT', 'AHM 10 MATTE', '11 7/8 X 8 3/4', 'DRAWER'],
        [3, 'MDF', 'VISTA FLAT', 'AHM 40', '11 7/8 X 8 3/4', 'DRAWER'],
        [2, 'MDF', 'SIERRA RAISED', 'AHM 10 MATTE', '11 7/8 X 9 11/16', 'DRAWER'],
        [2, 'MDF', 'VISTA FLAT', 'AHM  10 MATTE', '23 7/8 X 8 3/4', 'DRAWER'],
        [2, 'MDF', 'VISTA FLAT', 'AHM 10 MATTE', '29 7/8 X 8 3/4', 'DRAWER'],
        [2, 'MDF', 'VISTA FLAT', 'AHM 40', '29 7/8 X 8 3/4', 'DRAWER'],
        [1, 'MDF', 'SIERRA RAISED', 'AHM 10 MATTE', '35 7/8 X 6 15/16', 'DRAWER'],
        [1, 'MDF', 'SIERRA FLAT', 'AHM 20 MATTE', '23 7/8 X 2 3/4', 'MOULDING'],
        [1, 'MDF', 'VISTA FLAT', 'AHM 25', '29 7/8 X 2 3/4', 'MOULDING'],
        [1, 'MDF', 'SHAKER', 'AHM 50', '35 7/8 X 2 3/4', 'MOULDING'],
        [1, 'MDF', 'VISTA FLAT', 'AHM 80', '41 7/8 X 2 3/4', 'MOULDING'],
        [1, 'MDF', 'VISTA FLAT', 'AHM  10 MATTE', '59 7/8 X 2 3/4', 'MOULDING'],
        [1, 'MDF', 'VISTA FLAT', 'AHM 10 MATTE', '71 7/8 X 2 3/4', 'MOULDING'],
        [1, 'MDF', 'VISTA FLAT', 'AHM 40', '71 7/8 X 2 3/4', 'MOULDING'],
        [1, 'MDF', 'VISTA FLAT', 'AHM 10 MATTE`', '14 7/8 X 2 3/4', 'TOWER MOULDING']
    ]

    wb = openpyxl.Workbook()
    wb.create_sheet("PARTS DETAILS")
    wb.create_sheet("PARTS COLOR")
    ws = [ wb["PARTS DETAILS"], wb["PARTS COLOR"] ]

    write_parts_for_workshop(ws, items)

    wb.save('output.xlsx')

